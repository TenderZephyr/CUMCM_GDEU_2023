'''
Author: Jonty ljt20030312@Outlook.com
Date: 2023-09-09 22:33
LastEditTime: 2023-09-09 22:50
Description: 
'''
import pandas as pd
from openpyxl import Workbook

# 读取第一个Excel文件，包含分类信息
excel1 = pd.read_excel('D:\\桌面\\CUMCM\\topic\\附件1.xlsx')

# 读取第二个Excel文件，包含销售信息
excel2 = pd.read_excel('D:\\桌面\\CUMCM\\topic\\附件2.xlsx')

# 获取分类名称列表
categories = excel1['分类名称'].unique()

# 遍历每个分类
for category in categories:
    # 根据分类筛选销售数据
    filtered_data = excel2[excel2['单品编码'].isin(excel1[excel1['分类名称'] == category]['单品编码'])]
    
    # 创建一个新的Excel工作簿
    output_workbook = Workbook()
    
    # 创建一个新的工作表
    worksheet = output_workbook.active
    
    # 设置第一行为单品编码
    codes = filtered_data['单品编码'].unique()
    for col_num, code in enumerate(codes, start=2):
        worksheet.cell(row=1, column=col_num, value=code)
    
    # 设置第一列为销售日期
    dates = filtered_data['销售日期'].unique()
    for row_num, date in enumerate(dates, start=2):
        worksheet.cell(row=row_num, column=1, value=date)
    
    # 填充销量数据
    for i, date in enumerate(dates, start=2):
        for j, code in enumerate(codes, start=2):
            quantity = filtered_data[(filtered_data['销售日期'] == date) & (filtered_data['单品编码'] == code)]['销量(千克)'].sum()
            worksheet.cell(row=i, column=j, value=quantity)
    
    # 保存Excel文件，文件名使用分类名称
    output_workbook.save(f'D:\\桌面\\CUMCM\\classificationFile_sales\\productSales_day\\{category}.xlsx')

print("所有分类的数据已保存为单独的Excel文件")

